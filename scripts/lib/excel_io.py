# -*- coding: utf-8 -*-
"""Self-contained Excel helpers (no Flask / project Gateway)."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .constants import RESULT_SHEET_NAME

TITLE_HEADER_CANDIDATES = [
    "标题",
    "评论标题",
    "title",
    "review title",
    "headline",
    "subject",
]

CONTENT_HEADER_CANDIDATES = [
    "内容",
    "评论内容",
    "评论",
    "review content",
    "review text",
    "review",
    "body",
    "buyer comments",
    "customer reviews",
]


def _norm_header(value) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


def list_sheet_names(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def guess_review_sheet(sheet_names: list[str]) -> str | None:
    if not sheet_names:
        return None
    for name in sheet_names:
        low = name.lower()
        if any(k in low for k in ("评论", "review", "reviews", "feedback")):
            return name
    return sheet_names[0]


def read_headers(path: str | Path, sheet_name: str) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet 不存在：{sheet_name}")
        ws = wb[sheet_name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return []
        return ["" if c is None else str(c).strip() for c in first]
    finally:
        wb.close()


def _guess_column(headers: list[str], candidates: list[str], *, exclude: set[str] | None = None) -> str | None:
    exclude = exclude or set()
    normalized = [(_norm_header(h), h) for h in headers if str(h or "").strip() and h not in exclude]
    for cand in candidates:
        cand_n = _norm_header(cand)
        for nh, original in normalized:
            if nh == cand_n:
                return original
    for cand in candidates:
        cand_n = _norm_header(cand)
        for nh, original in normalized:
            if cand_n in nh or nh in cand_n:
                return original
    return None


def guess_title_column(headers: list[str]) -> str | None:
    return _guess_column(headers, TITLE_HEADER_CANDIDATES)


def guess_content_column(headers: list[str], *, title_column: str | None = None) -> str | None:
    exclude = {title_column} if title_column else set()
    return _guess_column(headers, CONTENT_HEADER_CANDIDATES, exclude=exclude)


def _cell_text(row, col_idx: int | None) -> str:
    if col_idx is None or row is None or col_idx >= len(row):
        return ""
    cell = row[col_idx]
    return "" if cell is None else str(cell).strip()


def build_review_text(title: str = "", content: str = "") -> str:
    title = (title or "").strip()
    content = (content or "").strip()
    if title and content:
        return f"标题：{title}\n内容：{content}"
    if title:
        return f"标题：{title}"
    if content:
        return f"内容：{content}"
    return ""


def format_ai_input_block(review: dict) -> str:
    rid = review.get("review_id") or ""
    review_text = (review.get("review_text") or "").strip()
    if not review_text:
        review_text = build_review_text(review.get("title") or "", review.get("content") or "")
    if not review_text:
        return ""
    return f"[review_id: {rid}]\n{review_text}"


def build_reviews_block(reviews: list[dict]) -> str:
    parts = [format_ai_input_block(item) for item in reviews]
    return "\n\n".join(p for p in parts if p)


def resolve_review_columns(
    path: str | Path,
    sheet_name: str | None = None,
    *,
    title_column: str | None = None,
    content_column: str | None = None,
) -> tuple[str, str | None, str | None]:
    sheets = list_sheet_names(path)
    if not sheets:
        raise ValueError("Excel 中没有任何 Sheet")
    sheet = (sheet_name or "").strip() or guess_review_sheet(sheets) or sheets[0]
    headers = read_headers(path, sheet)
    title = (title_column or "").strip() or guess_title_column(headers)
    content = (content_column or "").strip() or guess_content_column(headers, title_column=title)
    if not title and not content:
        raise ValueError("无法识别评论标题列或内容列")
    return sheet, title, content


def count_and_load_reviews(
    path: str | Path,
    sheet_name: str,
    content_column: str | None = None,
    title_column: str | None = None,
) -> list[dict]:
    title_column = (title_column or "").strip() or None
    content_column = (content_column or "").strip() or None
    if not title_column and not content_column:
        raise ValueError("请至少选择标题列或内容列之一")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet 不存在：{sheet_name}")
        ws = wb[sheet_name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return []
        headers = ["" if c is None else str(c).strip() for c in first]
        if title_column and title_column not in headers:
            raise ValueError(f"找不到标题列：{title_column}")
        if content_column and content_column not in headers:
            raise ValueError(f"找不到内容列：{content_column}")
        title_idx = headers.index(title_column) if title_column else None
        content_idx = headers.index(content_column) if content_column else None

        reviews = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            title = _cell_text(row, title_idx)
            content = _cell_text(row, content_idx)
            review_text = build_review_text(title, content)
            item = {
                "review_row": row_idx,
                "review_id": f"R{row_idx:06d}",
                "title": title,
                "content": content,
                "review_text": review_text,
                "skip_ai": not review_text,
            }
            item["ai_input"] = format_ai_input_block(item)
            reviews.append(item)
        return reviews
    finally:
        wb.close()


def write_analysis_workbook(
    source_path: str | Path,
    output_path: str | Path,
    *,
    summary_rows: list[dict],
) -> None:
    wb = load_workbook(source_path)
    for name in (RESULT_SHEET_NAME, "VOC分析结果", "VOC评论明细"):
        if name in wb.sheetnames:
            del wb[name]

    ws_sum = wb.create_sheet(RESULT_SHEET_NAME)
    ws_sum.append(["类型", "具体维度", "提及评论数", "提及频率", "代表性反馈"])
    for row in summary_rows:
        rate = float(row.get("mention_rate") or 0)
        feedback = row.get("representative_feedback") or row.get("core_description") or ""
        ws_sum.append(
            [
                row.get("item_type") or "",
                row.get("dimension") or "",
                int(row.get("mention_count") or 0),
                rate / 100.0,
                feedback,
            ]
        )
    for r in range(2, ws_sum.max_row + 1):
        ws_sum.cell(row=r, column=4).number_format = "0.00%"

    # Reviewability: freeze header + auto filter (sort already by 类型 → 提及数)
    ws_sum.freeze_panes = "A2"
    if ws_sum.max_row >= 1:
        ws_sum.auto_filter.ref = f"A1:E{ws_sum.max_row}"

    for idx, width in enumerate([12, 18, 14, 12, 56], start=1):
        ws_sum.column_dimensions[get_column_letter(idx)].width = width

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
